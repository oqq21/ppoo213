"""api_client.py — Mapleland.gg API 헬퍼"""
from __future__ import annotations
import re, time, requests
from your_app.common.io_cache import open_workbook_cached
from urllib.parse import urlencode
from typing import Dict, Any, List, Optional
from your_app.common.config import HEADERS, BASE_URL, ORDER, SHEET_TO_ITEMTYPE, dbg
from your_app.common.hangul_utils import to_choseong, is_choseong_query

_GENDER_MAP = {"남":"남자","남자":"남자","여":"여자","여자":"여자","공용":"공용"}
HAP_STATS_MAP = {"전사":"STRDEXACC","법사":"INTLUK","궁수":"STRDEX","도적":"DEXLUK","초보자":"","단도":"STRDEXLUK"}
_STAT_PREFIX = {"힘":"STR","덱":"DEX","인":"INT","럭":"LUK","공":"PAD","마":"MAD","hp":"MHP","HP":"MHP","피":"MHP",
                "이속":"Speed","점프":"Jump","명":"ACC","회피":"EVA","합마":"hapma"}
_ALLOWED_COLORS = {"purple","yellow","white","blue","gray"}
ZERO_OPTION_COMPONENTS = {
    "힘": ["힘"],
    "덱": ["덱"],
    "인": ["인", "마"],
    "럭": ["럭"],
    "공": ["공"],
    "마": ["인", "마"],
    "명": ["명"],
    "회": ["회피"],
    "회피": ["회피"],
    "물방": ["물방"],
    "마방": ["마방"],
    "이속": ["이속"],
    "점프": ["점프"],
    "피": ["HP", "피"],
    "HP": ["HP", "피"],
    "전사": ["힘", "덱", "명"],
    "궁수": ["힘", "덱"],
    "도적": ["덱", "럭"],
    "법사": ["인", "럭", "마"],
    "단도": ["힘", "덱", "럭"],
    "신점": ["힘", "덱", "명"],
    "신민": ["힘", "덱", "명"],
    "법지": ["인", "럭", "마"],
    "법행": ["인", "럭", "마"],
    "법신": ["인", "럭", "마"],
}

def normalize_gender(g) -> str:
    """0/1/2 → 남자/여자/생략, 한글 '남자/여자/공용' → '남자/여자/생략'."""
    if g is None:
        return ""
    s = str(g).strip()
    # 숫자 처리
    if s.replace(".", "", 1).isdigit():
        try:
            v = int(float(s))
            if v == 0:
                return "남자"
            if v == 1:
                return "여자"
            # 2 = 공용 → 생략
            return ""
        except Exception:
            return ""
    # 한글 처리
    mapped = _GENDER_MAP.get(s, "")
    return mapped if mapped in ("남자", "여자") else ""  # 공용은 생략



def infer_gender_from_name(name: str) -> str:
    m = re.search(r"\((남|여|공용)\)", str(name))
    return {"남":"남자","여":"여자","공용":"공용"}.get(m.group(1),"") if m else ""


def parse_zero_option_token(tok: str) -> List[str] | None:
    tok = (tok or "").strip()
    if not tok:
        return None
    m = re.fullmatch(r"([가-힣A-Za-z]+)\s*0", tok)
    if not m:
        return None
    return ZERO_OPTION_COMPONENTS.get(m.group(1))


def _normalize_option_parts(v) -> List[str]:
    if v is None:
        return []
    if isinstance(v, (list, tuple)):
        return [re.sub(r"\s+", "", str(x or "")) for x in v if str(x or "").strip()]
    s = str(v).strip()
    if not s:
        return []
    if s.startswith("[") and s.endswith("]"):
        s = s.strip("[]").replace("'", "").replace('"', "")
        return [re.sub(r"\s+", "", p) for p in s.split(",") if str(p).strip()]
    return [re.sub(r"\s+", "", s)]


def option_has_component(option_value, component: str) -> bool:
    component = str(component or "").strip()
    if not component:
        return False
    parts = _normalize_option_parts(option_value)
    if not parts:
        return False
    for part in parts:
        if component in ("피", "HP"):
            if part.upper().startswith("HP"):
                return True
        elif component == "마":
            if re.match(r"^마(?!방)", part):
                return True
        elif component == "회피":
            if part.startswith("회피") or part.startswith("회"):
                return True
        else:
            if part.startswith(component):
                return True
    return False


def option_has_any_component(option_value, components: List[str]) -> bool:
    return any(option_has_component(option_value, comp) for comp in (components or []))


def _find_item_code(sheet: str, item_name: str) -> str:
    """item.xlsx에서 sheet의 itemName으로 itemCode 조회. 실패 시 빈 문자열."""
    try:
        import openpyxl, os
        xls_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "..", "item.xlsx")
        xls_path = os.path.normpath(xls_path)
        wb = open_workbook_cached(xls_path, data_only=True)
        if sheet not in wb.sheetnames:
            return ""
        ws = wb[sheet]
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        idx_code = headers.index("itemCode")
        idx_name = headers.index("itemName")
        # exact match
        for row in ws.iter_rows(min_row=2, values_only=True):
            nm = str(row[idx_name]).strip() if row[idx_name] is not None else ""
            if nm == str(item_name).strip():
                return str(row[idx_code]).strip()
        # tolerant match (strip gender marks)
        def _norm(nm:str)->str:
            nm = str(nm or "")
            for t in ["(남)","(여)","남","여"]:
                nm = nm.replace(t,"")
            nm = re.sub(r"[^0-9a-zA-Z가-힣]+", "", nm)
            return nm.strip()
        target = _norm(item_name)
        for row in ws.iter_rows(min_row=2, values_only=True):
            nm = str(row[idx_name]).strip() if row[idx_name] is not None else ""
            if _norm(nm) == target:
                return str(row[idx_code]).strip()
        return ""
    except Exception:
        return ""
def parse_stat_token(tok: str) -> Dict[str, Any]:
    tok = (tok or "").strip()
    if not tok: return {}
    m_tuc = re.match(r"^(?:업|업횟)\s*(\d+)$", tok)
    if m_tuc:
        v = int(m_tuc.group(1)); return {"lowTuc": v, "highTuc": v}
    m_dando = re.match(r"단도\s*(\d+)$", tok)
    if m_dando:
        v = int(m_dando.group(1))
        if v == 0:
            return {}
        return {"hapStatsName":"STRDEXLUK","lowHapStatsValue":v,"highHapStatsValue":v}
    m_hap = re.match(r"(전사|법사|궁수|도적|초보자)\s*(\d+)$", tok)
    if m_hap:
        cls, v = m_hap.group(1), int(m_hap.group(2))
        if v == 0:
            return {}
        name = HAP_STATS_MAP.get(cls,""); 
        return ({"hapStatsName":name,"lowHapStatsValue":v,"highHapStatsValue":v} if name else {})
    # 신점/신민: 전사 n 처리 (서버 hap STRDEXACC)
    m_flag = re.match(r"(신점|신민)\s*(\d+)$", tok)
    if m_flag:
        v = int(m_flag.group(2))
        if v == 0:
            return {}
        return {"hapStatsName":"STRDEXACC","lowHapStatsValue":v,"highHapStatsValue":v}
    # 법지/법행: 법사 n 처리 (서버 hap INTLUK)
    m_b = re.match(r"(법지|법행)\s*(\d+)$", tok)
    if m_b:
        v = int(m_b.group(2))
        if v == 0:
            return {}
        return {"hapStatsName":"INTLUK","lowHapStatsValue":v,"highHapStatsValue":v}
    # 법신: API 파라미터 없음
    m_lps = re.match(r"법신\s*(\d+)$", tok)
    if m_lps:
        return {}
    m = re.match(r"([^\d]+)\s*(\d+)$", tok, flags=re.I)
    if not m: return {}
    key, v = m.group(1).strip(), int(m.group(2))
    pre = _STAT_PREFIX.get(key) or _STAT_PREFIX.get(key.lower()) or _STAT_PREFIX.get(key.upper())
    if pre and v == 0:
        return {}
    return ({f"lowinc{pre}": v, f"highinc{pre}": v} if pre else {})

def merge_stat_tokens(tokens: List[str]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for t in tokens or []: out.update(parse_stat_token(t))
    return out
def build_params(
    sheet: str,
    gender: str,
    reqlevel: int,
    item_name: str = "",
    stat_tokens: Optional[List[str]] = None,
    low_tuc: Optional[int] = None,
    job: Optional[str] = None,
) -> Dict[str, Any]:
    itype = SHEET_TO_ITEMTYPE.get(str(sheet).strip(), "")
    if not itype:
        print(f"[WARN] itemType 매핑 없음: {sheet}", flush=True)
        return {}

    g = normalize_gender(gender) or infer_gender_from_name(item_name)
    itype = SHEET_TO_ITEMTYPE.get(str(sheet).strip(), "")
    is_wand = (itype == "wand")
    is_staff = (itype == "staff")
    is_earring = (itype == "earrings")
    has_ma = any(re.match(r"^마\s*(\d+)$", str(t).strip()) for t in (stat_tokens or []))
    item_code = _find_item_code(str(sheet).strip(), str(item_name).strip()) if (is_wand or is_staff) else ""

    p: Dict[str, Any] = {
        "itemType": itype,
        "lowPrice": "",
        "highPrice": 9999999999,
        "lowLevel": int(reqlevel) if reqlevel else "",
        "highLevel": int(reqlevel) if reqlevel else "",
        "lowUpgrade": "",
        "highUpgrade": "",
    }
    if g:
        p["gender"] = g
    if isinstance(job, str) and job in {"전사","궁수","법사","도적"}:
        p["job"] = job
    if isinstance(low_tuc, int) and low_tuc >= 0:
        p["lowTuc"] = low_tuc
        p["highTuc"] = low_tuc

    p.update(merge_stat_tokens(stat_tokens or []))
    # 특수 규칙: 귀고리/스태프/완드 + '마 n' → MAD 대신 HAPMA 사용
    if (is_earring or is_wand or is_staff) and has_ma:
        v = None
        if "lowincMAD" in p and "highincMAD" in p:
            try:
                v = int(p.get("lowincMAD"))
            except Exception:
                v = None
            p.pop("lowincMAD", None); p.pop("highincMAD", None)
        if v is not None:
            p["lowHapma"] = v; p["highHapma"] = v
    # 스태프/완드: itemCode 사용 & gender 제거, itemType 제거
    if is_wand or is_staff:
        if item_code:
            p["itemCode"] = item_code
        p.pop("gender", None)
        p.pop("itemType", None)
    # 디버그 로그
    try:
        print(f"[DBG][API-PARAMS] sheet={sheet} rep={item_name} itype={itype} code={p.get('itemCode','')} has_ma={has_ma} hapma={p.get('lowHapma',None)}", flush=True)
    except Exception:
        pass
    return p





def build_url(params: dict) -> str:
    qs = []
    for k in ORDER:
        if k not in params: continue
        v = params[k]
        if k in ("lowPrice","highPrice") or v != "": qs.append((k,v))
    return BASE_URL + "?" + urlencode(qs, doseq=True)

def fetch_json_with_retries(params: Dict[str, Any], retries: int=3, delay: float=1.0):
    url = build_url(params)
    for i in range(1, retries+1):
        try:
            print(f"[CALL][{i}/{retries}] {url}", flush=True)
            r = requests.get(url, headers=HEADERS, timeout=10)
            print("[RESP]", r.status_code, flush=True)
            r.raise_for_status(); return r.json()
        except Exception as e:
            print(f"[ERR] try {i}/{retries} failed: {e}", flush=True)
            if i==retries: return None
            time.sleep(delay)

def _safe_get(d: dict, *keys, default=""):
    cur = d
    for k in keys:
        if not isinstance(cur, dict): return default
        cur = cur.get(k)
    return cur if cur is not None else default

def _extract_color(d: dict) -> str:
    io_, op = (d.get("itemOption") or {}), (d.get("tradeOption") or {})
    color = _safe_get(io_, "diff", "color", default="") or io_.get("color","") or op.get("color","") or d.get("color","") or ""
    color = str(color).strip().lower()
    return color if color in {"purple","yellow","white","blue","gray"} else "white"
def parse_trade_json(d: dict) -> dict:
    """API 항목 하나를 내부 표준 레코드로 변환."""
    if not isinstance(d, dict):
        return {}

    op = d.get("tradeOption") or {}
    io_ = d.get("itemOption") or {}
    tr = d.get("traderDiscordInfo") or {}

    # 판매자명
    seller_name = str(tr.get("global_name") or "").strip()

    # 거래 URL
    trade_url = f"https://mapleland.gg/trade/{d.get('url','')}" if d.get("url") else ""

    # 프로필 URL (traderDiscordInfo.id 사용)
    prof_id = str(tr.get("id") or "").strip()
    profile_url = f"https://mapleland.gg/profile/{prof_id}" if prof_id else ""

    # color
    diff = io_.get("diff") or {}
    color = str(diff.get("color") or "").strip().lower()

    # offer 원본(0/1/2) 그대로 보존
    offer_raw = op.get("offer", "")
    try:
        offer_raw = int(offer_raw)
    except Exception:
        pass  # 숫자가 아니면 그대로 둠

    return {
        "tradeType": d.get("tradeType"),
        "tradeStatus": d.get("tradeStatus"),
        "itemName": d.get("itemName"),
        "itemPrice": d.get("itemPrice"),
        "optionSummarize": io_.get("optionSummarize"),
        "comment": d.get("comment") or op.get("comment") or io_.get("comment") or "",
        "created_at": d.get("created_at"),
        "updated_at": d.get("updated_at"),
        "server": op.get("server"),
        "tuc": io_.get("tuc"),
        "color": color,
        "tradeUrl": trade_url,
        "profileUrl": profile_url,
        "global_name": seller_name,
        "offer_raw": offer_raw,   # ← UI에서 0/1/2 → 표시값으로 변환,
        "incSTR": int((io_ or {}).get("incSTR", 0) or 0),
        "incDEX": int((io_ or {}).get("incDEX", 0) or 0),
        "incINT": int((io_ or {}).get("incINT", 0) or 0),
        "incLUK": int((io_ or {}).get("incLUK", 0) or 0),
        "incPAD": int((io_ or {}).get("incPAD", 0) or 0),
        "incMAD": int((io_ or {}).get("incMAD", 0) or 0),
        "incACC": int((io_ or {}).get("incACC", 0) or 0),
        "incJump": int((io_ or {}).get("incJump", 0) or 0),
        "incMHP": int((io_ or {}).get("incMHP", 0) or 0),
        "STRDEXLUK": int((io_ or {}).get("STRDEXLUK", 0) or 0),
    }


def filter_records_by_query(records: List[dict], query: str) -> List[dict]:
    if not query or not records: return records
    q = query.replace(" ",""); out = []
    if is_choseong_query(q):
        for r in records:
            if q in to_choseong(str(r.get("itemName",""))): out.append(r)
    else:
        for r in records:
            if q in str(r.get("itemName","")).replace(" ",""): out.append(r)
    dbg(f"post-filter: {len(records)} -> {len(out)} by '{q}'")
    return out
