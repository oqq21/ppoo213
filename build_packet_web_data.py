from __future__ import annotations

import argparse
import json
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook


KST = timezone(timedelta(hours=9))
ROOT = Path(__file__).resolve().parent
WORKSPACE_ROOT = ROOT.parents[1]
RESOURCE_DIR = WORKSPACE_ROOT / "자료"
DATA_DIR = WORKSPACE_ROOT / "data"

STAT_INDEXES = {
    "upgrade_left": 0,
    "work_count": 1,
    "STR": 2,
    "DEX": 3,
    "INT": 4,
    "LUK": 5,
    "HP": 6,
    "MP": 7,
    "PAD": 8,
    "MAD": 9,
    "PDD": 10,
    "MDD": 11,
    "ACC": 12,
    "EVA": 13,
    "SPEED": 14,
    "JUMP": 16,
}

ITEMP_COLUMNS = {
    "힘": "STR",
    "덱": "DEX",
    "인": "INT",
    "럭": "LUK",
    "공": "PAD",
    "마": "MAD",
    "물방": "PDD",
    "마방": "MDD",
    "명": "ACC",
    "회피": "EVA",
    "이속": "SPEED",
    "점프": "JUMP",
    "HP": "HP",
    "MP": "MP",
}

GEM_STAT_SUFFIXES = {
    0: ("다이아몬드", "PAD"),
    1: ("사파이어", "MAD"),
    2: ("가넷", "ACC"),
    3: ("오팔", "EVA"),
    4: ("자수정", "SPEED"),
    5: ("아쿠아마린", "JUMP"),
    6: ("토파즈", "HP"),
    7: ("에메랄드", "MP"),
    8: ("힘의 크리스탈", "STR"),
    9: ("지혜의 크리스탈", "INT"),
    10: ("행운의 크리스탈", "LUK"),
    11: ("민첩성의 크리스탈", "DEX"),
}

GEM_GRADES = {
    6_000: ("하급", 750_000),
    16_000: ("중급", 1_500_000),
    26_000: ("상급", 2_250_000),
}

GEM_OPTION_LABELS = {
    0: ("공", (1, 2, 3)),
    1: ("마", (1, 2, 3)),
    2: ("명", (2, 3, 5)),
    3: ("회", (2, 3, 5)),
    4: ("이속", (2, 3, 5)),
    5: ("점프", (1, 2, 3)),
    6: ("HP", (10, 20, 30)),
    7: ("MP", (10, 20, 30)),
    8: ("힘", (2, 3, 5)),
    9: ("인", (2, 3, 5)),
    10: ("럭", (2, 3, 5)),
    11: ("덱", (1, 3, 5)),
}

GEM_BASE_FEE = 2_250_000
GEM_VALUE_RATE = 0.9
WEAPON_SHEETS = {
    "한손검", "한손도끼", "한손둔기", "단검", "두손검", "두손도끼",
    "두손둔기", "창", "폴암", "활", "석궁", "아대", "너클", "총",
}


def safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError, OverflowError):
        try:
            return int(float(value))
        except (TypeError, ValueError, OverflowError):
            return default


def game_time_to_kst(value: Any) -> str:
    raw = safe_int(value)
    if raw <= 60_000_000_000_000:
        return ""
    try:
        return (datetime(1, 1, 1) + timedelta(milliseconds=raw, hours=9)).isoformat(sep=" ")
    except (OverflowError, ValueError):
        return ""


def packet_time_to_kst(value: Any) -> str:
    try:
        return datetime.fromtimestamp(float(value), KST).replace(tzinfo=None).isoformat(sep=" ")
    except (TypeError, ValueError, OverflowError, OSError):
        return ""


def load_item_catalog(path: Path) -> tuple[dict[int, dict[str, Any]], dict[str, pd.DataFrame]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    catalog: dict[int, dict[str, Any]] = {}
    web_sheets: dict[str, pd.DataFrame] = {}

    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration:
            continue
        indexes = {str(value).strip(): index for index, value in enumerate(headers) if value is not None}
        if "itemCode" not in indexes or "itemName" not in indexes:
            continue

        output_rows = []
        output_headers = [
            "itemCode", "itemName", "gender", "reqLevel", "category",
            "공", "마", "힘", "덱", "인", "럭", "명",
            "물방", "마방", "이속", "점프", "회피", "직업",
            "HP", "MP",
        ]
        for row in rows:
            code = row[indexes["itemCode"]] if indexes["itemCode"] < len(row) else None
            if not isinstance(code, int) or not 1_000_000 <= code < 2_000_000:
                continue
            name = row[indexes["itemName"]] if indexes["itemName"] < len(row) else None
            if not name:
                continue

            base = {key: 0 for key in set(STAT_INDEXES) - {"upgrade_left", "work_count"}}
            for source_name, target_name in ITEMP_COLUMNS.items():
                index = indexes.get(source_name)
                if index is not None and index < len(row):
                    base[target_name] = safe_int(row[index])
            catalog[code] = {
                "item_name": str(name).strip(),
                "sheet": worksheet.title,
                "base": base,
                "attack_gem_warning": worksheet.title in WEAPON_SHEETS,
            }

            output = []
            for header in output_headers:
                index = indexes.get(header)
                output.append(row[index] if index is not None and index < len(row) else None)
            output_rows.append(output)

        if output_rows:
            web_sheets[worksheet.title] = pd.DataFrame(output_rows, columns=output_headers)

    workbook.close()
    return catalog, web_sheets


def gem_option_code_from_item_code(item_code: int) -> int | None:
    offset = item_code - 4_250_000
    if offset < 0:
        return None
    stat_suffix, grade_index = divmod(offset, 100)
    if stat_suffix not in GEM_STAT_SUFFIXES or grade_index not in (0, 1, 2):
        return None
    return (6_000, 16_000, 26_000)[grade_index] + stat_suffix


def load_gem_prices(path: Path) -> dict[int, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    prices: dict[int, int] = {}
    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration:
            continue
        code_columns = [
            index for index, value in enumerate(headers)
            if str(value or "").strip().lower().startswith("itemcode")
        ]
        for row in rows:
            for code_index in code_columns:
                code = safe_int(row[code_index] if code_index < len(row) else None, -1)
                option_code = gem_option_code_from_item_code(code)
                if option_code is None:
                    continue
                price = safe_int(row[code_index + 1] if code_index + 1 < len(row) else None, -1)
                if price >= 0:
                    prices[option_code] = price
    workbook.close()

    expected = {
        base + suffix
        for base in GEM_GRADES
        for suffix in GEM_STAT_SUFFIXES
    }
    missing = sorted(expected - set(prices))
    if missing:
        raise ValueError(f"보석 평균가 누락: {missing}")
    return prices


def iter_gem_codes(value: Any) -> Iterable[int]:
    values = value if isinstance(value, list) else [value]
    for item in values:
        code = safe_int(item, -1)
        for base in GEM_GRADES:
            if base <= code <= base + 11:
                yield code
                break


def gem_details(option_codes: Any, prices: dict[int, int]) -> tuple[str, int, int]:
    codes = list(iter_gem_codes(option_codes))
    if not codes:
        return "", 0, 0

    labels = []
    total_cost = GEM_BASE_FEE
    for code in codes:
        grade_base = max(base for base in GEM_GRADES if base <= code)
        _grade_name, processing_fee = GEM_GRADES[grade_base]
        suffix = code - grade_base
        grade_index = tuple(GEM_GRADES).index(grade_base)
        stat_name, values = GEM_OPTION_LABELS[suffix]
        labels.append(f"{stat_name}+{values[grade_index]}")
        total_cost += prices[code] + processing_fee
    recognized = round(total_cost * GEM_VALUE_RATE)
    return ", ".join(labels), total_cost, recognized


def gem_stat_totals(option_codes: Any) -> dict[str, int]:
    result = {stat: 0 for stat in set(STAT_INDEXES) - {"upgrade_left", "work_count"}}
    for code in iter_gem_codes(option_codes):
        grade_base = max(base for base in GEM_GRADES if base <= code)
        grade_index = tuple(GEM_GRADES).index(grade_base)
        suffix = code - grade_base
        _name, stat = GEM_STAT_SUFFIXES[suffix]
        _label, values = GEM_OPTION_LABELS[suffix]
        result[stat] += values[grade_index]
    return result


def parse_equipment_rows(
    db_path: Path,
    status: str,
    catalog: dict[int, dict[str, Any]],
    gem_prices: dict[int, int],
) -> list[dict[str, Any]]:
    completed = status == "completed"
    type_index = 2 if completed else 1
    connection = sqlite3.connect(db_path.resolve().as_uri() + "?mode=ro", uri=True)
    cutoff = (datetime.now(KST) - timedelta(days=60)).timestamp()
    cursor = connection.execute(
        f"""
        SELECT packet_time, raw_row_json
        FROM items
        WHERE CAST(json_extract(raw_row_json, '$[{type_index}]') AS INTEGER) = 1
          AND packet_time >= ?
        """,
        (cutoff,),
    )

    records = []
    for packet_time, raw_json in cursor:
        try:
            raw = json.loads(raw_json)
        except (TypeError, json.JSONDecodeError):
            continue
        if not isinstance(raw, list) or len(raw) < 14:
            continue

        if completed:
            code, quantity = safe_int(raw[3], -1), safe_int(raw[4])
            total_price, unit_price = safe_int(raw[9]), safe_int(raw[10])
            internal_time = game_time_to_kst(raw[1])
            event_time = packet_time_to_kst(packet_time)
        else:
            code, quantity = safe_int(raw[2], -1), safe_int(raw[3])
            total_price, unit_price = safe_int(raw[8]), safe_int(raw[9])
            internal_time = game_time_to_kst(raw[10])
            event_time = packet_time_to_kst(packet_time)

        item = catalog.get(code)
        if item is None:
            continue
        stats = raw[11] if isinstance(raw[11], list) else []
        options = raw[13] if isinstance(raw[13], list) else []
        gem_text, gem_cost, recognized_gem_value = gem_details(options, gem_prices)
        gem_stats = gem_stat_totals(options)
        recognized_codes = list(iter_gem_codes(options))
        record = {
            "status": status,
            "captured_at": packet_time_to_kst(packet_time),
            # active=만료 예정시간, completed=판매완료 시간.
            # 화면에는 표시하지 않고 active↔completed 3일 판정에만 사용한다.
            "internal_time": internal_time,
            "event_time": event_time,
            "itemCode": code,
            "itemName": item["item_name"],
            "sheet": item["sheet"],
            "quantity": quantity,
            "total_price": total_price,
            "unit_price": unit_price,
            "gem_options": gem_text,
            "gem_cost": gem_cost,
            "recognized_gem_value": recognized_gem_value,
            "true_price": max(0, unit_price - recognized_gem_value),
            "option_codes": ",".join(
                str(safe_int(code))
                for code in options
                if safe_int(code) != 0
            ),
            "attack_gem_warning": bool(item["attack_gem_warning"]),
            "gem_cell_style": (
                "red"
                if item["attack_gem_warning"] and 26_000 not in recognized_codes
                else "green"
                if len(recognized_codes) == 3
                and all(26_000 <= code <= 26_011 for code in recognized_codes)
                else "white"
            ),
        }
        for stat, value in gem_stats.items():
            record[f"gem_{stat}"] = int(value)
        for name, index in STAT_INDEXES.items():
            record[f"total_{name}"] = safe_int(stats[index]) if index < len(stats) else 0
        records.append(record)
    connection.close()
    return records


def infer_missing_base_hp(
    catalog: dict[int, dict[str, Any]],
    frames: Iterable[pd.DataFrame],
) -> dict[int, int]:
    counts: dict[int, Counter[int]] = defaultdict(Counter)
    for frame in frames:
        if frame.empty:
            continue
        clean = frame[frame["total_work_count"].eq(0)]
        for code, hp in clean[["itemCode", "total_HP"]].itertuples(index=False):
            counts[int(code)][int(hp)] += 1

    supplemented = {}
    for code, values in counts.items():
        item = catalog.get(code)
        if item is None or safe_int(item["base"].get("HP")) != 0 or not values:
            continue
        hp = max(values, key=lambda value: (values[value], -value))
        if hp > 0:
            item["base"]["HP"] = hp
            supplemented[code] = hp
    return supplemented


def add_base_and_additional_stats(
    frame: pd.DataFrame,
    catalog: dict[int, dict[str, Any]],
) -> pd.DataFrame:
    if frame.empty:
        return frame
    result = frame.copy()
    stat_names = set(STAT_INDEXES) - {"upgrade_left", "work_count"}
    for stat in stat_names:
        base_map = {code: safe_int(item["base"].get(stat)) for code, item in catalog.items()}
        result[f"base_{stat}"] = result["itemCode"].map(base_map).fillna(0).astype("int32")
        result[f"add_{stat}"] = (
            result[f"total_{stat}"].astype("int64") - result[f"base_{stat}"].astype("int64")
        ).astype("int32")
    return result


def write_site_item_file(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="패킷 SQLite를 Streamlit용 장비 parquet로 변환합니다.")
    parser.add_argument("--itemp", type=Path, default=RESOURCE_DIR / "itemp.xlsx")
    parser.add_argument(
        "--gem-prices",
        type=Path,
        default=RESOURCE_DIR / "주문서_광석_판매통계.xlsx",
    )
    parser.add_argument(
        "--active-db",
        type=Path,
        default=DATA_DIR / "store" / "market_active.sqlite",
    )
    parser.add_argument(
        "--completed-db",
        type=Path,
        default=DATA_DIR / "store" / "market_completed.sqlite",
    )
    parser.add_argument("--output-dir", type=Path, default=ROOT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    catalog, web_sheets = load_item_catalog(args.itemp)
    prices = load_gem_prices(args.gem_prices)
    print(f"[LOAD] 장비 itemCode {len(catalog):,}개 / 보석 가격 {len(prices):,}종")

    active = pd.DataFrame(parse_equipment_rows(args.active_db, "active", catalog, prices))
    completed = pd.DataFrame(parse_equipment_rows(args.completed_db, "completed", catalog, prices))
    print(f"[LOAD] active {len(active):,}건 / completed {len(completed):,}건")

    supplemented = infer_missing_base_hp(catalog, [active, completed])
    if supplemented:
        print(f"[HP] itemp 누락 기본 HP 보완 {len(supplemented):,}개")
    active = add_base_and_additional_stats(active, catalog)
    completed = add_base_and_additional_stats(completed, catalog)

    active_path = args.output_dir / "packet_active.parquet"
    completed_path = args.output_dir / "packet_completed.parquet"
    item_path = args.output_dir / "item.xlsx"
    gem_path = args.output_dir / "gem_prices.json"
    active.to_parquet(active_path, engine="pyarrow", compression="zstd", index=False)
    completed.to_parquet(completed_path, engine="pyarrow", compression="zstd", index=False)
    write_site_item_file(item_path, web_sheets)
    gem_path.write_text(
        json.dumps(
            {
                str(code): int(price)
                for code, price in sorted(prices.items())
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"[OK] {active_path} ({active_path.stat().st_size / 1024 / 1024:.1f} MB)")
    print(f"[OK] {completed_path} ({completed_path.stat().st_size / 1024 / 1024:.1f} MB)")
    print(f"[OK] {item_path}")
    print(f"[OK] {gem_path}")


if __name__ == "__main__":
    main()
